from openpyxl import Workbook, load_workbook

# Crear y escribir en un archivo Excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'Nombre'
ws['B1'] = 'Edad'
ws.append(['Ana', 23])
ws.append(['Luis', 30])
wb.save('archivo.xlsx')

# Leer el archivo Excel
wb = load_workbook('archivo.xlsx')
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
